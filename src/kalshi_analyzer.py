#!/usr/bin/env python3
"""
Kalshi Market Analyzer
======================
Pulls live prediction market data from Kalshi's public API,
identifies markets closing within specified time windows,
displays prices, and flags wide bid-ask spreads.

Author: Jose Hernandez
Assignment: Winworks Take-Home
"""

import httpx
import pendulum
import argparse
import json
from rich.console import Console
from rich.table import Table
from rich.panel import Panel
from rich import box
from dataclasses import dataclass, asdict
from typing import Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Kalshi public API base URL
KALSHI_API_BASE = "https://api.elections.kalshi.com/trade-api/v2"

# Threshold for "wide" spread (10% = 0.10)
WIDE_SPREAD_THRESHOLD = 0.10

console = Console()


@dataclass
class Market:
    """Represents a Kalshi prediction market."""
    ticker: str
    title: str
    subtitle: str
    yes_bid: float
    yes_ask: float
    no_bid: float
    no_ask: float
    close_time: pendulum.DateTime
    volume: int
    status: str
    category: str

    @property
    def spread(self) -> float:
        """Calculate bid-ask spread as a decimal."""
        return self.yes_ask - self.yes_bid

    @property
    def spread_percent(self) -> float:
        """Spread as percentage."""
        return self.spread * 100

    @property
    def is_wide_spread(self) -> bool:
        """Check if spread exceeds threshold."""
        return self.spread >= WIDE_SPREAD_THRESHOLD

    @property
    def hours_until_close(self) -> float:
        """Hours remaining until market closes."""
        now = pendulum.now('UTC')
        diff = self.close_time.diff(now)
        return diff.in_hours() + (diff.in_minutes() % 60) / 60

    @property
    def time_until_close_str(self) -> str:
        """Human-readable time until close."""
        hours = self.hours_until_close
        if hours < 0:
            return "CLOSED"
        elif hours < 1:
            return f"{int(hours * 60)}m"
        elif hours < 24:
            return f"{hours:.1f}h"
        elif hours < 168:  # 7 days
            return f"{hours/24:.1f}d"
        else:
            return f"{hours/24/7:.1f}w"

    @property
    def midpoint(self) -> float:
        """Calculate midpoint price (fair value estimate)."""
        if self.yes_bid and self.yes_ask:
            return (self.yes_bid + self.yes_ask) / 2
        return 0.5


def fetch_markets(limit: int = 500, cursor: str = None) -> list[dict]:
    """
    Fetch markets from Kalshi public API.

    Args:
        limit: Maximum number of markets to fetch
        cursor: Pagination cursor

    Returns:
        List of market data dictionaries
    """
    console.print("[cyan]Fetching markets from Kalshi API...[/cyan]")

    url = f"{KALSHI_API_BASE}/markets"
    params = {
        "limit": limit,
        "status": "open",
    }
    if cursor:
        params["cursor"] = cursor

    try:
        with httpx.Client(timeout=30.0) as client:
            response = client.get(url, params=params)
            response.raise_for_status()
            data = response.json()

            markets = data.get("markets", [])
            console.print(f"[green]✓ Fetched {len(markets)} markets[/green]")
            return markets

    except httpx.HTTPStatusError as e:
        console.print(f"[red]HTTP Error: {e.response.status_code}[/red]")
        raise
    except httpx.RequestError as e:
        console.print(f"[red]Request Error: {e}[/red]")
        raise


def parse_market(raw: dict) -> Optional[Market]:
    """
    Parse raw API response into Market object.

    Args:
        raw: Raw market dictionary from API

    Returns:
        Market object or None if parsing fails
    """
    try:
        # Use expected_expiration_time (when market resolves) instead of close_time (when trading ends)
        # This gives us the actual resolution time which is more useful for analysis
        close_time_str = raw.get("expected_expiration_time") or raw.get("close_time")
        if not close_time_str:
            return None

        close_time = pendulum.parse(close_time_str)

        # Extract prices (in cents, convert to dollars)
        yes_bid = raw.get("yes_bid", 0) / 100 if raw.get("yes_bid") else 0
        yes_ask = raw.get("yes_ask", 0) / 100 if raw.get("yes_ask") else 0
        no_bid = raw.get("no_bid", 0) / 100 if raw.get("no_bid") else 0
        no_ask = raw.get("no_ask", 0) / 100 if raw.get("no_ask") else 0

        return Market(
            ticker=raw.get("ticker", "N/A"),
            title=raw.get("title", "Unknown"),
            subtitle=raw.get("subtitle", ""),
            yes_bid=yes_bid,
            yes_ask=yes_ask,
            no_bid=no_bid,
            no_ask=no_ask,
            close_time=close_time,
            volume=raw.get("volume", 0),
            status=raw.get("status", "unknown"),
            category=raw.get("category", ""),
        )
    except Exception as e:
        return None


def filter_closing_within(markets: list[Market], hours: int = 24) -> list[Market]:
    """
    Filter markets closing within specified hours.

    Args:
        markets: List of Market objects
        hours: Time window in hours (default 24)

    Returns:
        Filtered list of markets
    """
    now = pendulum.now('UTC')
    cutoff = now.add(hours=hours)

    closing_soon = [
        m for m in markets
        if now < m.close_time <= cutoff
    ]

    # Sort by close time (soonest first)
    closing_soon.sort(key=lambda m: m.close_time)

    return closing_soon


def filter_by_spread(markets: list[Market], min_spread: float = 0.0) -> list[Market]:
    """Filter markets with spread above minimum."""
    return [m for m in markets if m.spread >= min_spread]


def filter_by_volume(markets: list[Market], min_volume: int = 0) -> list[Market]:
    """Filter markets with volume above minimum."""
    return [m for m in markets if m.volume >= min_volume]


def display_markets_table(markets: list[Market], title: str):
    """
    Display markets in a formatted table.

    Args:
        markets: List of Market objects to display
        title: Table title
    """
    if not markets:
        console.print(Panel(f"[yellow]No markets found for: {title}[/yellow]"))
        return

    console.print(f"\n[bold cyan]{title}[/bold cyan]")
    console.print(f"[dim]Found {len(markets)} markets[/dim]\n")

    # Create table with explicit width control
    table = Table(
        box=box.SIMPLE,
        show_header=True,
        header_style="bold white",
        padding=(0, 1),
    )

    # Add columns - simplified for better display
    table.add_column("Market", width=50, no_wrap=True)
    table.add_column("Yes", justify="right", width=6)
    table.add_column("No", justify="right", width=6)
    table.add_column("Spread", justify="right", width=8)
    table.add_column("Expires", justify="right", width=8)
    table.add_column("Vol", justify="right", width=8)

    # Add rows
    for market in markets[:15]:  # Limit to 15 rows
        # Determine spread color
        spread_style = "red bold" if market.is_wide_spread else "green"

        # Truncate title if needed
        display_title = market.title[:48] + ".." if len(market.title) > 50 else market.title

        # Format prices as cents (more intuitive for prediction markets)
        yes_cents = f"{int(market.yes_bid * 100)}¢" if market.yes_bid else "—"
        no_cents = f"{int(market.no_bid * 100)}¢" if market.no_bid else "—"

        # Add wide flag to spread column
        spread_text = f"{market.spread_percent:.0f}%"
        if market.is_wide_spread:
            spread_text = f"[red bold]{spread_text}⚠[/red bold]"
        else:
            spread_text = f"[green]{spread_text}[/green]"

        table.add_row(
            display_title,
            yes_cents,
            no_cents,
            spread_text,
            market.time_until_close_str,
            f"{market.volume:,}",
        )

    console.print(table)
    if len(markets) > 15:
        console.print(f"[dim]... and {len(markets) - 15} more markets[/dim]")


def display_wide_spreads(markets: list[Market]):
    """
    Highlight markets with wide bid-ask spreads.

    Args:
        markets: List of Market objects
    """
    wide_spread_markets = [m for m in markets if m.is_wide_spread]
    wide_spread_markets.sort(key=lambda m: m.spread, reverse=True)

    if not wide_spread_markets:
        console.print("\n[green]✓ No markets with wide spreads (>10%) found.[/green]")
        return

    console.print(f"\n[bold red]⚠️  WIDE SPREAD ALERT: {len(wide_spread_markets)} markets[/bold red]")
    console.print("[dim]Markets with bid-ask spread > 10% (potential liquidity issues)[/dim]\n")

    table = Table(box=box.SIMPLE, show_header=True, header_style="bold red")

    table.add_column("Market", width=45)
    table.add_column("Spread", justify="right", width=8)
    table.add_column("Bid→Ask", justify="center", width=15)
    table.add_column("Exp", justify="right", width=6)

    for market in wide_spread_markets[:10]:
        table.add_row(
            market.title[:43] + ".." if len(market.title) > 45 else market.title,
            f"[red bold]{market.spread_percent:.0f}%[/red bold]",
            f"{int(market.yes_bid*100)}¢ → {int(market.yes_ask*100)}¢",
            market.time_until_close_str,
        )

    console.print(table)


def display_high_volume(markets: list[Market], top_n: int = 10):
    """Display highest volume markets."""
    sorted_markets = sorted(markets, key=lambda m: m.volume, reverse=True)[:top_n]

    if not sorted_markets:
        return

    console.print(f"\n[bold magenta]📊 TOP {top_n} HIGHEST VOLUME MARKETS[/bold magenta]\n")

    table = Table(box=box.SIMPLE, show_header=True, header_style="bold magenta")
    table.add_column("Rank", justify="right", width=4)
    table.add_column("Market", width=45)
    table.add_column("Volume", justify="right")
    table.add_column("Yes $", justify="right")
    table.add_column("Spread", justify="right")

    for i, market in enumerate(sorted_markets, 1):
        spread_style = "red" if market.is_wide_spread else "green"
        table.add_row(
            f"#{i}",
            market.title[:43] + ".." if len(market.title) > 45 else market.title,
            f"[bold]{market.volume:,}[/bold]",
            f"${market.yes_bid:.2f}",
            f"[{spread_style}]{market.spread_percent:.1f}%[/{spread_style}]",
        )

    console.print(table)


def display_summary(all_markets: list[Market], time_windows: dict[str, list[Market]]):
    """
    Display summary statistics.

    Args:
        all_markets: All fetched markets
        time_windows: Dict of time window name to filtered markets
    """
    console.print("\n")
    summary_table = Table(title="📈 Summary Statistics", box=box.ROUNDED, show_header=True)
    summary_table.add_column("Metric", style="cyan")
    summary_table.add_column("Value", justify="right", style="white")

    summary_table.add_row("Total Markets Analyzed", f"[green]{len(all_markets):,}[/green]")

    for window_name, markets in time_windows.items():
        wide_count = sum(1 for m in markets if m.is_wide_spread)
        summary_table.add_row(
            f"Closing in {window_name}",
            f"[yellow]{len(markets):,}[/yellow] ({wide_count} wide spread)"
        )

    # Overall stats
    total_volume = sum(m.volume for m in all_markets)
    avg_spread = sum(m.spread_percent for m in all_markets) / len(all_markets) if all_markets else 0
    wide_total = sum(1 for m in all_markets if m.is_wide_spread)

    summary_table.add_row("Total Volume (all markets)", f"[magenta]{total_volume:,}[/magenta] contracts")
    summary_table.add_row("Average Spread", f"[cyan]{avg_spread:.2f}%[/cyan]")
    summary_table.add_row("Wide Spread Markets (>10%)", f"[red]{wide_total}[/red]")

    console.print(summary_table)


def export_to_json(markets: list[Market], filename: str):
    """Export markets to JSON file."""
    data = []
    for m in markets:
        data.append({
            "ticker": m.ticker,
            "title": m.title,
            "yes_bid": m.yes_bid,
            "yes_ask": m.yes_ask,
            "no_bid": m.no_bid,
            "no_ask": m.no_ask,
            "spread_percent": m.spread_percent,
            "is_wide_spread": m.is_wide_spread,
            "close_time": m.close_time.isoformat(),
            "hours_until_close": m.hours_until_close,
            "volume": m.volume,
        })

    with open(filename, 'w') as f:
        json.dump(data, f, indent=2)

    console.print(f"[green]✓ Exported {len(data)} markets to {filename}[/green]")


def export_to_excel(markets: list[Market], wide_spread_markets: list[Market], filename: str):
    """
    Export markets to Excel file with formatting.

    Creates two sheets:
    - All Markets: Complete market data
    - Wide Spreads: Markets with spread > 10% (highlighted)
    """
    wb = Workbook()

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    wide_spread_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
    money_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Sheet 1: All Markets
    ws1 = wb.active
    ws1.title = "All Markets"

    headers = ["Ticker", "Market", "Yes Bid", "Yes Ask", "No Bid", "No Ask",
               "Spread %", "Wide Spread", "Expires In", "Hours Left", "Volume"]

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Write data
    for row_num, market in enumerate(markets, 2):
        data_row = [
            market.ticker,
            market.title[:60] + "..." if len(market.title) > 60 else market.title,
            f"${market.yes_bid:.2f}",
            f"${market.yes_ask:.2f}",
            f"${market.no_bid:.2f}",
            f"${market.no_ask:.2f}",
            f"{market.spread_percent:.1f}%",
            "YES" if market.is_wide_spread else "NO",
            market.time_until_close_str,
            round(market.hours_until_close, 1),
            market.volume,
        ]

        for col, value in enumerate(data_row, 1):
            cell = ws1.cell(row=row_num, column=col, value=value)
            cell.border = thin_border

            # Highlight wide spread rows
            if market.is_wide_spread:
                cell.fill = wide_spread_fill

            # Highlight money columns
            if col in [3, 4, 5, 6]:
                cell.alignment = Alignment(horizontal='right')

    # Adjust column widths
    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 50
    ws1.column_dimensions['C'].width = 10
    ws1.column_dimensions['D'].width = 10
    ws1.column_dimensions['E'].width = 10
    ws1.column_dimensions['F'].width = 10
    ws1.column_dimensions['G'].width = 10
    ws1.column_dimensions['H'].width = 12
    ws1.column_dimensions['I'].width = 12
    ws1.column_dimensions['J'].width = 10
    ws1.column_dimensions['K'].width = 10

    # Sheet 2: Wide Spreads Only
    ws2 = wb.create_sheet(title="Wide Spreads Alert")

    headers2 = ["Ticker", "Market", "Spread %", "Bid → Ask", "Expires", "Volume"]

    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = PatternFill(start_color="C62828", end_color="C62828", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for row_num, market in enumerate(wide_spread_markets, 2):
        data_row = [
            market.ticker,
            market.title[:50] + "..." if len(market.title) > 50 else market.title,
            f"{market.spread_percent:.1f}%",
            f"${market.yes_bid:.2f} → ${market.yes_ask:.2f}",
            market.time_until_close_str,
            market.volume,
        ]

        for col, value in enumerate(data_row, 1):
            cell = ws2.cell(row=row_num, column=col, value=value)
            cell.border = thin_border
            cell.fill = wide_spread_fill

    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 45
    ws2.column_dimensions['C'].width = 10
    ws2.column_dimensions['D'].width = 20
    ws2.column_dimensions['E'].width = 10
    ws2.column_dimensions['F'].width = 10

    # Sheet 3: Summary Statistics
    ws3 = wb.create_sheet(title="Summary")

    summary_data = [
        ["Metric", "Value"],
        ["Total Markets Analyzed", len(markets)],
        ["Wide Spread Markets (>10%)", len(wide_spread_markets)],
        ["Average Spread", f"{sum(m.spread_percent for m in markets) / len(markets):.2f}%" if markets else "0%"],
        ["Total Volume", sum(m.volume for m in markets)],
        ["Report Generated", pendulum.now().format("YYYY-MM-DD HH:mm:ss UTC")],
    ]

    for row_num, row_data in enumerate(summary_data, 1):
        for col, value in enumerate(row_data, 1):
            cell = ws3.cell(row=row_num, column=col, value=value)
            cell.border = thin_border
            if row_num == 1:
                cell.font = header_font
                cell.fill = header_fill

    ws3.column_dimensions['A'].width = 30
    ws3.column_dimensions['B'].width = 25

    # Save
    wb.save(filename)
    console.print(f"[green]✓ Exported to Excel: {filename}[/green]")
    console.print(f"[dim]  - Sheet 1: All Markets ({len(markets)} rows)[/dim]")
    console.print(f"[dim]  - Sheet 2: Wide Spreads ({len(wide_spread_markets)} rows)[/dim]")
    console.print(f"[dim]  - Sheet 3: Summary Statistics[/dim]")


def main():
    """Main entry point."""
    parser = argparse.ArgumentParser(description="Analyze Kalshi prediction markets")
    parser.add_argument("--hours", type=int, default=24, help="Time window in hours (default: 24)")
    parser.add_argument("--limit", type=int, default=500, help="Max markets to fetch (default: 500)")
    parser.add_argument("--min-volume", type=int, default=0, help="Minimum volume filter")
    parser.add_argument("--wide-only", action="store_true", help="Show only wide spread markets")
    parser.add_argument("--json", type=str, metavar="FILE", help="Export results to JSON file")
    parser.add_argument("--excel", type=str, metavar="FILE", help="Export results to Excel file (.xlsx)")
    args = parser.parse_args()

    console.print(Panel.fit(
        "[bold cyan]🎯 Kalshi Market Analyzer[/bold cyan]\n"
        "[dim]Real-time prediction market analysis[/dim]",
        border_style="cyan",
    ))
    console.print()

    try:
        # Step 1: Fetch markets
        raw_markets = fetch_markets(limit=args.limit)

        # Step 2: Parse markets
        console.print("[cyan]Processing market data...[/cyan]")
        markets = []
        for raw in raw_markets:
            market = parse_market(raw)
            if market:
                markets.append(market)

        console.print(f"[green]✓ Parsed {len(markets)} valid markets[/green]\n")

        # Apply volume filter if specified
        if args.min_volume > 0:
            markets = filter_by_volume(markets, args.min_volume)
            console.print(f"[yellow]Filtered to {len(markets)} markets with volume >= {args.min_volume}[/yellow]\n")

        # Step 3: Analyze different time windows
        time_windows = {
            "24 hours": filter_closing_within(markets, hours=24),
            "48 hours": filter_closing_within(markets, hours=48),
            "7 days": filter_closing_within(markets, hours=168),
        }

        # Step 4: Display results
        # Primary view: Markets closing in specified time window
        primary_markets = filter_closing_within(markets, hours=args.hours)
        display_markets_table(primary_markets, f"Markets Closing in {args.hours} Hours")

        # If no markets in 24h, show 7-day window
        if len(primary_markets) == 0 and args.hours == 24:
            console.print("\n[yellow]No markets closing in 24h. Showing 7-day window:[/yellow]")
            display_markets_table(time_windows["7 days"][:15], "Markets Closing in 7 Days")

        # Step 5: Show wide spread alerts
        display_wide_spreads(markets)

        # Step 6: Show high volume markets
        display_high_volume(markets)

        # Step 7: Show summary
        display_summary(markets, time_windows)

        # Step 8: Export to JSON if requested
        if args.json:
            export_to_json(primary_markets if primary_markets else markets, args.json)

        # Step 9: Export to Excel if requested
        if args.excel:
            wide_spread_list = [m for m in markets if m.is_wide_spread]
            wide_spread_list.sort(key=lambda m: m.spread, reverse=True)
            export_to_excel(
                primary_markets if primary_markets else markets,
                wide_spread_list,
                args.excel
            )

        console.print("\n[green bold]✓ Analysis complete![/green bold]")
        console.print("[dim]Data fetched from Kalshi Public API at " +
                     pendulum.now().format("YYYY-MM-DD HH:mm:ss") + " UTC[/dim]\n")

    except Exception as e:
        console.print(f"[red bold]Error: {e}[/red bold]")
        import traceback
        traceback.print_exc()
        raise


if __name__ == "__main__":
    main()
